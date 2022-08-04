#!/usr/bin/env python
# coding: utf-8

# In[1]:


import requests
import pandas as pd
from openpyxl import load_workbook
from copy import copy
import codecs
 
answerFileName = 'test.xlsx'
templateFileName = 'kek.xlsx'
 
htmlFileName = ""
 
class Lesson(object):
    def __init__(self):
        self.dayOfTheWeak = ""
        self.time = ""
        self.lessonName = ""
        self.frequency = ""
        self.group = ""
        self.type = ""
        self.classNumber = ""

class Person(object):
    def __init__(self, name):
        self.name = name
        self.lessons = []

        
table = pd.read_html('2.html', encoding="utf8")[1]
#print(table[0])
persons = []
numbersOfRowsWith1Column = 0
numbersOfRowsWithSomeColumns = 0
lastPerson = Person("name")
lastDayOfTheWeak = ""
 
wb = load_workbook(templateFileName)
ws = wb['Лист1']
wb.save(answerFileName) # проверяем можно ли сохранить файл, если исключение, то нужно закрыть excel с этим файлом
 
# парсим из html в список persons с lessons
for rowIndex in range(0, len(table[0])):
    if table[0][rowIndex] == table[1][rowIndex]:
        numbersOfRowsWith1Column += 1
        numbersOfRowsWithSomeColumns = 0
    else:
        if numbersOfRowsWith1Column == 1:
            lastDayOfTheWeak = table[0][rowIndex - 1]
        if numbersOfRowsWith1Column == 2:
            lastPerson = Person(table[0][rowIndex - 2])
            persons.append(lastPerson)
            lastDayOfTheWeak = table[0][rowIndex - 1]
        numbersOfRowsWith1Column = 0
 
        numbersOfRowsWithSomeColumns += 1
        if numbersOfRowsWithSomeColumns > 1:
            newLesson = Lesson()
            newLesson.time = table[0][rowIndex]
            newLesson.frequency = table[1][rowIndex]
            newLesson.classNumber = table[2][rowIndex]
            newLesson.group = table[3][rowIndex]
            newLesson.lessonName = table[4][rowIndex]
            newLesson.type = table[5][rowIndex]
            newLesson.dayOfTheWeak = lastDayOfTheWeak
            lastPerson.lessons.append(newLesson)

# метод позволяет разбить ячейки в зависимости от частоты занятий: Еженедельно, Числитель, Знаменатель
def merged_cells(row, column, sheet, typeMerged):
    cell1 = sheet.cell(row, column)
    cell2 = sheet.cell(row + 1, column)
    unmergedAll = False
    for mergedCell in sheet.merged_cells.ranges:
        if cell1.coordinate in mergedCell and cell2.coordinate in mergedCell:
            if typeMerged == 1:
                return
            else:
                unmergedAll = True
                sheet.unmerge_cells(start_row=row, start_column=column, end_row=row + 1, end_column=column + 1)
 
    if not unmergedAll:
        cell1 = sheet.cell(row, column)
        cell2 = sheet.cell(row, column + 1)
        for mergedCell in sheet.merged_cells.ranges:
            if cell1.coordinate in mergedCell and cell2.coordinate in mergedCell:
                sheet.unmerge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
                break
 
        cell1 = sheet.cell(row + 1, column)
        cell2 = sheet.cell(row + 1, column + 1)
        for mergedCell in sheet.merged_cells.ranges:
            if cell1.coordinate in mergedCell and cell2.coordinate in mergedCell:
                sheet.unmerge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
                break
 
    if typeMerged == 1:
        sheet.merge_cells(start_row=row, start_column=column, end_row=row + 1, end_column=column + 1)
    else:
        sheet.cell(row + 1, column).fill = copy(sheet.cell(row, column).fill)
        sheet.cell(row + 1, column).font = copy(sheet.cell(row, column).font)
        sheet.cell(row + 1, column).border = copy(sheet.cell(row, column).border)
        sheet.merge_cells(start_row=row, start_column=column, end_row=row, end_column=column + 1)
        sheet.merge_cells(start_row=row + 1, start_column=column, end_row=row + 1, end_column=column + 1)
 
 
# сохраняем в excel файл
for person in persons:
    #номер колонки с нужным преподователем
    columnNumberPerson = -1
    for column in range(1, ws.max_column):
        if not ws.cell(row=3, column=column).value is None and person.name in ws.cell(row=3, column=column).value:
            columnNumberPerson = column
 
    if columnNumberPerson == -1:
        personNameWithoutIO = person.name
        for i in personNameWithoutIO:
            if not i.isalpha() and i != ' ':
                personNameWithoutIO = personNameWithoutIO.replace(i, '')
 
        personNameWithoutIO = personNameWithoutIO.split()[0].strip() + ' '
        for column in range(1, ws.max_column):
            if not ws.cell(row=3, column=column).value is None and personNameWithoutIO in ws.cell(row=3, column=column).value:
                columnNumberPerson = column
 
    if columnNumberPerson == -1:
        #print(person.name + " NOT FOUND")
        continue
 
    # сохраняем пары для преподователся
    for lesson in person.lessons:
        rowNumberDayOfWeak = -1
        # находим строку с денм недели
        for rowNumber in range(1, ws.max_row):
            rowValue = ws.cell(row=rowNumber, column=1).value
            if not rowValue is None:
                rowValue = rowValue.replace('\n', "")
                if rowValue.lower() == lesson.dayOfTheWeak.replace('\n', "").lower():
                    rowNumberDayOfWeak = rowNumber
 
        if rowNumberDayOfWeak == -1:
            print("DAY NOT FOUND")
 
        # находим строку со временем начала занятий
        rowNumberTime = -1
        for rowNumber in range(rowNumberDayOfWeak, ws.max_row):
            rowValue = ws.cell(row=rowNumber, column=2).value
            if not rowValue is None:
                if rowValue[:5] == lesson.time[:5]:
                    rowNumberTime = rowNumber
                    break
 
        if rowNumberTime == -1:
            print("TIME NOT FOUND")
 
        # разбиваем или объединяем ячейки
        if lesson.frequency == "Еженедельно":
            merged_cells(rowNumberTime, columnNumberPerson, ws, 1)
        else:
            merged_cells(rowNumberTime, columnNumberPerson, ws, 2)
 
 
        if lesson.frequency == "Знаменатель":
            rowNumberTime += 1
 
        ws.cell(row=rowNumberTime, column=columnNumberPerson).value = lesson.lessonName + "\n" + lesson.group
        if not lesson.classNumber is None:
            ws.cell(row=rowNumberTime, column=columnNumberPerson).value = ws.cell(row=rowNumberTime,
                                                                                  column=columnNumberPerson).value + ", ауд. " + lesson.classNumber
 
 
wb.save(templateFileName)

